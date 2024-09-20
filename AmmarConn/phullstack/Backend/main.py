import fitz  # PyMuPDF
import pytesseract
from PIL import Image
import io
import os
import pandas as pd
from docx import Document as WordDocument
from openpyxl import load_workbook
from langchain_community.embeddings import OllamaEmbeddings
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_community.vectorstores import Chroma
from langchain.chains import ConversationalRetrievalChain
from langchain_community.chat_message_histories import ChatMessageHistory
from langchain.memory import ConversationBufferMemory
from langchain_groq import ChatGroq
from dotenv import load_dotenv
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
import uvicorn
from pydantic import BaseModel
import logging

# Loading environment variables from .env file
load_dotenv()

app = FastAPI()

logging.basicConfig(level=logging.INFO)

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allows all origins
    allow_credentials=True,
    allow_methods=["*"],  # Allows all methods
    allow_headers=["*"],  # Allows all headers
)

# Initialize conversation chain with GROQ language model
groq_api_key = os.environ['GROQ_API_KEY']

llm_groq = ChatGroq(
    groq_api_key=groq_api_key, model_name="llama-3.1-70b-versatile",
    temperature=0.2
)

# Global variable to store the conversation chain
global_chain = None

def extract_text_from_image(image_bytes):
    """
    Extract text from an image using OCR.
    """
    image = Image.open(io.BytesIO(image_bytes))
    ocr_text = pytesseract.image_to_string(image)
    return ocr_text

def process_files(files):
    texts = []
    metadatas = []
    images_save_path = 'imgg/'  # Path to save images

    # Ensure the images directory exists
    if not os.path.exists(images_save_path):
        os.makedirs(images_save_path)

    for file in files:
        file_content = file.file.read()
        file_ext = os.path.splitext(file.filename)[1].lower()
        logging.info(f"Processing file: {file.filename} with extension: {file_ext}")

        try:
            if file_ext in ['.pdf']:
                # Process PDF files
                pdf = fitz.open(stream=file_content, filetype="pdf")
                for page_num in range(len(pdf)):
                    page = pdf[page_num]
                    text = page.get_text()
                    if text:
                        texts.append(text)

                    images = page.get_images(full=True)
                    for img_index, img in enumerate(images, start=1):
                        xref = img[0]
                        base_image = pdf.extract_image(xref)
                        image_bytes = base_image["image"]

                        # Direct OCR extraction without saving the image
                        ocr_text = extract_text_from_image(image_bytes)
                        if ocr_text.strip():
                            texts.append(ocr_text)

            elif file_ext in ['.csv']:
                # Process CSV files
                df = pd.read_csv(io.StringIO(file_content.decode()))
                texts.append(df.to_string())

            elif file_ext in ['.txt']:
                # Process TXT files
                texts.append(file_content.decode())

            elif file_ext in ['.docx']:
                # Process Word files
                doc = WordDocument(io.BytesIO(file_content))
                text = "\n".join([para.text for para in doc.paragraphs])
                texts.append(text)

            elif file_ext in ['.xlsx']:
                # Process Excel files
                wb = load_workbook(io.BytesIO(file_content), data_only=True)
                text = ""
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for row in ws.iter_rows(values_only=True):
                        text += " ".join([str(cell) for cell in row]) + "\n"
                texts.append(text)

            elif file_ext in ['.jpeg', '.jpg', '.png']:
                # Process image files (JPEG, JPG, PNG)
                ocr_text = extract_text_from_image(file_content)
                if ocr_text.strip():
                    texts.append(ocr_text)

            else:
                logging.warning(f"Unsupported file type: {file_ext}")

        except Exception as e:
            logging.error(f"Error processing file {file.filename}: {e}")
            raise

    # Aggregate and split texts into chunks
    aggregated_text = " ".join(texts)
    text_splitter = RecursiveCharacterTextSplitter(chunk_size=1200, chunk_overlap=50)
    file_texts = text_splitter.split_text(aggregated_text)

    # Populate metadata for each chunk
    if file_texts:
        metadatas = [{"source": f"{i}-{file.filename}"} for i in range(len(file_texts))]

    if len(file_texts) == 0:
        raise ValueError("No texts extracted from the provided files.")

    # Create a Chroma vector store
    embeddings = OllamaEmbeddings(model="nomic-embed-text")
    docsearch = Chroma.from_texts(file_texts, embeddings, metadatas=metadatas)

    # Initialize message history for conversation
    message_history = ChatMessageHistory()

    # Memory for conversational context
    memory = ConversationBufferMemory(
        memory_key="chat_history",
        output_key="answer",
        chat_memory=message_history,
        return_messages=True,
    )

    # Create a chain that uses the Chroma vector store
    chain = ConversationalRetrievalChain.from_llm(
        llm=ChatGroq(),
        chain_type="stuff",
        retriever=docsearch.as_retriever(),
        memory=memory,
        return_source_documents=True,
    )

    return chain

@app.post("/upload/")
async def upload_file(files: list[UploadFile] = File(...)):
    global global_chain
    try:
        logging.info(f"Received {len(files)} files for processing.")
        global_chain = process_files(files)
        logging.info("Files processed successfully.")
        return {"message": f"Successfully processed {len(files)} files."}
    except Exception as e:
        logging.error(f"Error processing files: {e}")
        raise HTTPException(status_code=500, detail=str(e))

class ChatRequest(BaseModel):
    message: str

@app.post("/chat/")
async def chat(request: ChatRequest):
    global global_chain
    if not global_chain:
        raise HTTPException(status_code=400, detail="No files have been processed. Please upload files first.")

    prompt = (
        "You are a friendly and professional university chatbot. "
        "Your tone should be polite and engaging, suitable for a teenage audience. "
        "You should provide helpful responses related to the university documents and general inquiries. "
        "Feel free to offer casual conversation while staying within a professional and respectful tone."
        "Answer to the question in a clear and concise manner"
        "Also respond to general greetings and notes be a bit emotional and conversational"
        "Respond to Thanking and similar statements with welcoming notes"
    )

    context = f"{prompt}\nUser: {request.message}\nBot:"

    try:
        res = global_chain.invoke({"question": context})
        answer = res["answer"]
        source_documents = res["source_documents"]

        if source_documents:
            source_names = [f"source_{i}" for i in range(len(source_documents))]
            answer += f"\nSources: {', '.join(source_names)}"

        return {"response": answer}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)