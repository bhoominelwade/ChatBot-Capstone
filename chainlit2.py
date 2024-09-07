import fitz  # PyMuPDF
import pytesseract
from PIL import Image
import io
import os
import PyPDF2
import pandas as pd
from docx import Document as WordDocument
from openpyxl import load_workbook
from langchain_community.embeddings import OllamaEmbeddings
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_community.vectorstores import Chroma
from langchain.chains import ConversationalRetrievalChain
from langchain_community.chat_message_histories import ChatMessageHistory
from langchain.memory import ConversationBufferMemory
from langchain_groq import ChatGroq
from dotenv import load_dotenv

# Loading environment variables from .env file
load_dotenv()

# Initialize conversation chain with GROQ language model
groq_api_key = os.environ['GROQ_API_KEY']

llm_groq = ChatGroq(
    groq_api_key=groq_api_key, model_name="llama-3.1-70b-versatile",
    temperature=0.2
)

def extract_text_from_image(image_bytes):
    image = Image.open(io.BytesIO(image_bytes))
    ocr_text = pytesseract.image_to_string(image)
    return ocr_text

def process_files(files):
    texts = []
    metadatas = []
    images_save_path = 'imgg/'  # Path to save images

    # Ensure the images directory exists
    if not os.path.exists(images_save_path):
        os.makedirs(images_save_path)

    for file_path in files:
        file_ext = os.path.splitext(file_path)[1].lower()

        if file_ext in ['.pdf']:
            # Process PDF files
            pdf = fitz.open(file_path)
            for page_num in range(len(pdf)):
                page = pdf[page_num]
                text = page.get_text()
                if text:
                    texts.append(text)

                images = page.get_images(full=True)
                for img_index, img in enumerate(images, start=1):
                    xref = img[0]
                    base_image = pdf.extract_image(xref)
                    image_bytes = base_image["image"]
                    image_ext = base_image["ext"]

                    image_name = f"page_{page_num+1}_img_{img_index}.{image_ext}"
                    image_path = os.path.join(images_save_path, image_name)
                    with open(image_path, 'wb') as img_file:
                        img_file.write(image_bytes)

                    ocr_text = extract_text_from_image(image_bytes)
                    if ocr_text.strip():
                        texts.append(ocr_text)

        elif file_ext in ['.csv']:
            # Process CSV files
            df = pd.read_csv(file_path)
            texts.append(df.to_string())

        elif file_ext in ['.txt']:
            # Process TXT files
            with open(file_path, 'r') as file:
                texts.append(file.read())

        elif file_ext in ['.docx']:
            # Process Word files
            doc = WordDocument(file_path)
            text = "\n".join([para.text for para in doc.paragraphs])
            texts.append(text)

        elif file_ext in ['.xlsx']:
            # Process Excel files
            wb = load_workbook(file_path, data_only=True)
            text = ""
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    text += " ".join([str(cell) for cell in row]) + "\n"
            texts.append(text)

        else:
            print(f"Unsupported file type: {file_ext}")

        # Create metadata for each chunk
        text_splitter = RecursiveCharacterTextSplitter(chunk_size=1200, chunk_overlap=50)
        file_texts = text_splitter.split_text(" ".join(texts))
        texts.extend(file_texts)
        file_metadatas = [{"source": f"{i}-{os.path.basename(file_path)}"} for i in range(len(file_texts))]
        metadatas.extend(file_metadatas)

    # Create a Chroma vector store
    embeddings = OllamaEmbeddings(model="nomic-embed-text")
    docsearch = Chroma.from_texts(texts, embeddings, metadatas=metadatas)

    # Initialize message history for conversation
    message_history = ChatMessageHistory()

    # Memory for conversational context
    memory = ConversationBufferMemory(
        memory_key="chat_history",
        output_key="answer",
        chat_memory=message_history,
        return_messages=True,
    )

    # Create a chain that uses the Chroma vector store
    chain = ConversationalRetrievalChain.from_llm(
        llm=llm_groq,
        chain_type="stuff",
        retriever=docsearch.as_retriever(),
        memory=memory,
        return_source_documents=True,
    )

    return chain

def main():
    # Get the list of files to process
    files = input("Please provide the paths of files (comma-separated): ").split(',')

    # Process the files and initialize the conversational chain
    chain = process_files([file.strip() for file in files])

    print(f"Processing {len(files)} files done. You can now ask questions!")

    # Interactive loop to ask questions
    while True:
        user_input = input("You: ")
        if user_input.lower() in ['exit', 'quit']:
            print("Goodbye!")
            break

        res = chain.invoke({"question": user_input})
        answer = res["answer"]
        source_documents = res["source_documents"]

        # Process source documents if available
        if source_documents:
            source_names = [f"source_{i}" for i in range(len(source_documents))]
            answer += f"\nSources: {', '.join(source_names)}"

        print(f"Bot: {answer}")

if __name__ == "__main__":
    main()