# Required libraries for handling PDFs, images, and OCR
import fitz  # PyMuPDF for working with PDF files
import pytesseract  # OCR tool for extracting text from images
from PIL import Image  # Image processing library
import io  # For handling byte streams
import os  # For operating system dependent functionality

# Libraries for data manipulation and file handling
import pandas as pd  # Data analysis and manipulation library
from docx import Document as WordDocument  # For handling Word documents
from openpyxl import load_workbook  # For reading and writing Excel files
from datetime import datetime, timedelta
import mimetypes
from typing import Optional
import uuid
from firebase_admin import firestore

# LangChain specific libraries for embedding, text splitting, and vector storage
from langchain_community.embeddings import OllamaEmbeddings  # For embeddings
from langchain.text_splitter import RecursiveCharacterTextSplitter  # For splitting text
from langchain_community.vectorstores import Chroma  # For vector storage
from langchain.chains import ConversationalRetrievalChain  # For conversational chains
from langchain_community.chat_message_histories import ChatMessageHistory  # For message history management
from langchain.memory import ConversationBufferMemory  # For conversation memory

# For working with Groq API in LangChain
from langchain_groq import ChatGroq  # For integrating with Groq services

# Environment variable management
from dotenv import load_dotenv  # For loading environment variables from .env file

# FastAPI for building the API
from fastapi import FastAPI, UploadFile, File, HTTPException, Form  # For creating the FastAPI app and handling file uploads
from fastapi.middleware.cors import CORSMiddleware  # type: ignore # For handling CORS

# Uvicorn for serving the FastAPI app
import uvicorn  # ASGI server for running FastAPI applications

# Pydantic for data validation and settings management
from pydantic import BaseModel  # For defining data models

# Logging for application logging
import logging  # For logging messages and errors

# Firebase imports
import firebase_admin
from firebase_admin import credentials, storage, firestore
import datetime
from firebase_admin import storage

from tenacity import retry, stop_after_attempt, wait_exponential
from fuzzywuzzy import process

# Initialize Firebase
cred = credentials.Certificate(r"Tera")
firebase_admin.initialize_app(cred, {
    'storageBucket': 'capstone-4eff9.appspot.com'
})

# Get a reference to the storage service and Firestore
bucket = storage.bucket()
db = firestore.client()

# Loading environment variables from .env file
load_dotenv()

app = FastAPI()

logging.basicConfig(level=logging.INFO)

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Initialize conversation chain with GROQ language model
groq_api_key = os.environ['GROQ_API_KEY']

llm_groq = ChatGroq(
    groq_api_key=groq_api_key,
    model_name="llama-3.1-70b-versatile",
    temperature=0.2
)

# Global variable to store the conversation chain
global_chain = None

def extract_text_from_image(image_bytes):
    """
    Extract text from an image using OCR.
    """
    image = Image.open(io.BytesIO(image_bytes))
    ocr_text = pytesseract.image_to_string(image)
    return ocr_text

def normalize_role(role):
    """Normalize role names to lowercase singular form."""
    role = role.lower()
    if role.endswith('s'):
        role = role[:-1]  # Remove trailing 's'
    return role

def upload_to_firebase(file_content, filename, role):
    """
    Upload a file to Firebase Storage with metadata and correct MIME type.
    """
    blob = bucket.blob(filename)
    
    # Detect MIME type
    content_type, _ = mimetypes.guess_type(filename)
    if content_type is None:
        content_type = 'application/octet-stream'  # Default to binary if type can't be guessed
    
    blob.upload_from_string(file_content, content_type=content_type)
    
    # Normalize and set custom metadata
    normalized_role = normalize_role(role)
    metadata = {'role': normalized_role}
    blob.metadata = metadata
    blob.patch()

    # Verify that metadata was set correctly
    updated_blob = bucket.get_blob(filename)
    if updated_blob.metadata != metadata:
        logging.error(f"Failed to set metadata for {filename}. Expected {metadata}, got {updated_blob.metadata}")
    else:
        logging.info(f"File uploaded to Firebase with metadata: {filename}, role: {normalized_role}, content_type: {content_type}")

    return blob.public_url

def process_files(files, role):
    texts = []
    metadatas = []

    for file in files:
        file_content = file.file.read()
        file_ext = os.path.splitext(file.filename)[1].lower()
        logging.info(f"Processing file: {file.filename} with extension: {file_ext}")

        try:
            # Upload file to Firebase Storage with role metadata
            file_url = upload_to_firebase(file_content, file.filename, role)
            logging.info(f"File uploaded to Firebase: {file_url}")

            if file_ext in ['.pdf']:
                # Process PDF files
                pdf = fitz.open(stream=file_content, filetype="pdf")
                for page_num in range(len(pdf)):
                    page = pdf[page_num]
                    text = page.get_text()
                    if text:
                        texts.append(text)

                    images = page.get_images(full=True)
                    for img_index, img in enumerate(images, start=1):
                        xref = img[0]
                        base_image = pdf.extract_image(xref)
                        image_bytes = base_image["image"]

                        # OCR extraction
                        ocr_text = extract_text_from_image(image_bytes)
                        if ocr_text.strip():
                            texts.append(ocr_text)

            elif file_ext in ['.csv']:
                # Process CSV files
                df = pd.read_csv(io.StringIO(file_content.decode()))
                texts.append(df.to_string())

            elif file_ext in ['.txt']:
                # Process TXT files
                texts.append(file_content.decode())

            elif file_ext in ['.docx']:
                # Process Word files
                doc = WordDocument(io.BytesIO(file_content))
                text = "\n".join([para.text for para in doc.paragraphs])
                texts.append(text)

            elif file_ext in ['.xlsx']:
                # Process Excel files
                wb = load_workbook(io.BytesIO(file_content), data_only=True)
                text = ""
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for row in ws.iter_rows(values_only=True):
                        text += " ".join([str(cell) for cell in row]) + "\n"
                texts.append(text)

            elif file_ext in ['.jpeg', '.jpg', '.png']:
                # Process image files (JPEG, JPG, PNG)
                ocr_text = extract_text_from_image(file_content)
                if ocr_text.strip():
                    texts.append(ocr_text)

            else:
                logging.warning(f"Unsupported file type: {file_ext}")

        except Exception as e:
            logging.error(f"Error processing file {file.filename}: {e}")
            raise

    # Aggregate and split texts into chunks
    aggregated_text = " ".join(texts)
    text_splitter = RecursiveCharacterTextSplitter(chunk_size=1200, chunk_overlap=50)
    file_texts = text_splitter.split_text(aggregated_text)

    # Populate metadata for each chunk
    if file_texts:
        metadatas = [{"source": f"{i}-{file.filename}", "role": role} for i in range(len(file_texts))]

    if len(file_texts) == 0:
        raise ValueError("No texts extracted from the provided files.")

    # Create a Chroma vector store
    embeddings = OllamaEmbeddings(model="nomic-embed-text")
    docsearch = Chroma.from_texts(file_texts, embeddings, metadatas=metadatas)

    # Initialize message history for conversation
    message_history = ChatMessageHistory()

    # Memory for conversational context
    memory = ConversationBufferMemory(
        memory_key="chat_history",
        output_key="answer",
        chat_memory=message_history,
        return_messages=True,
    )

    # Create a chain that uses the Chroma vector store
    chain = ConversationalRetrievalChain.from_llm(
        llm=ChatGroq(),
        chain_type="stuff",
        retriever=docsearch.as_retriever(),
        memory=memory,
        return_source_documents=True,
    )

    return chain

def can_access_document(user_role, document_role):
    """
    Check if a user with given role can access a document with specified role.
    Implements role hierarchy: HOD/Dean > Teacher > Student
    """
    # Normalize roles to lowercase
    user_role = user_role.lower().replace('/', '_').replace(' ', '_')
    document_role = document_role.lower().replace('/', '_').replace(' ', '_')
    
    # Define role hierarchy
    role_hierarchy = {
        'hod_dean': ['hod_dean', 'teacher', 'student'],
        'teacher': ['teacher', 'student'],
        'student': ['student']
    }
    
    # Check if user's role exists in hierarchy
    if user_role not in role_hierarchy:
        logging.error(f"Invalid user role: {user_role}")
        return False
    
    # Check if document role is accessible to user
    return document_role in role_hierarchy[user_role]

@app.post("/upload/")
async def upload_file(files: list[UploadFile] = File(...), role: str = Form(...)):
    global global_chain
    try:
        normalized_role = normalize_role(role)
        logging.info(f"Received {len(files)} files for processing with normalized role: {normalized_role}")
        global_chain = process_files(files, normalized_role)
        logging.info("Files processed successfully.")
        return {"message": f"Successfully processed {len(files)} files for role: {normalized_role}"}
    except Exception as e:
        logging.error(f"Error processing files: {e}")
        raise HTTPException(status_code=500, detail=str(e))

class ChatRequest(BaseModel):
    message: str

@app.post("/retrieve-document/")
async def retrieve_document(request: dict):
    file_name = request.get("fileName")
    user_role = normalize_role(request.get("userRole", ""))

    logging.info(f"Document retrieval request - File: {file_name}, User Role: {user_role}")

    if not file_name or not user_role:
        logging.error("Missing fileName or userRole in request")
        raise HTTPException(status_code=400, detail="Missing fileName or userRole")

    try:
        # Get list of all blobs in the bucket
        blobs = list(bucket.list_blobs())
        blob_names = [blob.name for blob in blobs]

        # Find the best matching file
        best_match, score = process.extractOne(file_name, blob_names)
        if score < 80:
            logging.error(f"No close match found for: {file_name}")
            raise HTTPException(status_code=404, detail=f"Document not found: {file_name}")

        blob = bucket.blob(best_match)
        
        # Get document metadata
        blob.reload()  # Ensure we have the latest metadata
        document_role = blob.metadata.get('role', 'student') if blob.metadata else 'student'
        
        logging.info(f"Document metadata - File Role: {document_role}")
        
        # Check access permission
        if not can_access_document(user_role, document_role):
            logging.warning(f"Access denied - User Role: {user_role}, Document Role: {document_role}")
            return {
                "canAccess": False,
                "message": "You don't have permission to access this document."
            }

        # Generate signed URL for authorized access
        signed_url = blob.generate_signed_url(
            version="v4",
            expiration=datetime.timedelta(minutes=15),
            method="GET"
        )
        
        logging.info(f"Access granted - Generated signed URL for {best_match}")
        return {
            "canAccess": True,
            "downloadUrl": signed_url,
            "fileName": best_match
        }

    except Exception as e:
        logging.exception(f"Error in retrieve_document: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")
    
@app.post("/chat/")
async def chat(request: ChatRequest):
    global global_chain
    if not global_chain:
        raise HTTPException(status_code=400, detail="No files have been processed. Please upload files first.")

    prompt = (
    "You are a friendly and professional university chatbot. "
    "Your tone should be polite and engaging, suitable for a teenage audience. "
    "You should provide helpful responses related to the university documents and general inquiries. "
    "Feel free to offer casual conversation while staying within a professional and respectful tone. "
    "Answer to the question in a clear and concise manner. "
    "Also respond to general greetings and notes in a friendly, conversational tone. "
    "Respond to thanking and similar statements with welcoming notes. "
    )

    context = f"{prompt}\nUser: {request.message}\nBot:"

    try:
        res = global_chain.invoke({"question": context})
        answer = res["answer"]
        source_documents = res["source_documents"]

        if source_documents:
            source_names = [f"source_{i}" for i in range(len(source_documents))]
            answer += f"\nSources: {', '.join(source_names)}"

        return {"response": answer}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    

@app.get("/announcements/{role}")
async def get_announcements(role: str):
    """
    Get announcements based on user role.
    Implements role hierarchy: HOD/Dean can see all, Teachers see Teacher and Student, Students see only Student
    """
    try:
        # Get Firestore database reference
        db = firestore.client()
        
        # Normalize role (handle plural forms)
        role = role.lower()
        if role == "students":
            role = "student"
        elif role == "teachers":
            role = "teacher"
        elif role in ["hod/dean", "hod_dean", "hod/deans"]:
            role = "hod_dean"
        
        # Define role hierarchy
        role_access = {
            'hod_dean': ['hod_dean', 'teacher', 'student'],
            'teacher': ['teacher', 'student'],
            'student': ['student']
        }
        
        if role not in role_access:
            raise HTTPException(status_code=400, detail=f"Invalid role specified: {role}")
            
        # Query announcements based on role hierarchy
        announcements_ref = db.collection('announcements')
        
        # Create query
        docs = announcements_ref.order_by('timestamp', direction=firestore.Query.DESCENDING).get()
        
        # Filter announcements based on role hierarchy
        announcements = []
        for doc in docs:
            data = doc.to_dict()
            data['id'] = doc.id
            # Normalize the announcement role as well
            announcement_role = data.get('role', '').lower()
            if announcement_role.endswith('s'):
                announcement_role = announcement_role[:-1]
            # Only add announcements if the role is in the access list
            if announcement_role in role_access[role]:
                # Convert timestamp to string for JSON serialization
                if data.get('timestamp'):
                    utc_time = data['timestamp']
                    ist_time = utc_time + datetime.timedelta(hours=5, minutes=30)  # Convert to IST
                    data['timestamp'] = ist_time.strftime('%Y-%m-%d %I:%M %p')
                announcements.append(data)  # Add the announcement to the list
            
        return {"announcements": announcements}
    except Exception as e:
        logging.error(f"Error fetching announcements: {e}")
        raise HTTPException(status_code=500, detail=f"Error fetching announcements: {str(e)}")

@app.delete("/announcement/{announcement_id}")
async def delete_announcement(announcement_id: str):
    """Delete an announcement by ID."""
    try:
        # Get Firestore database reference
        db = firestore.client()
        
        # Delete the announcement document
        db.collection('announcements').document(announcement_id).delete()
        
        return {"message": "Announcement deleted successfully", "status": "success"}
    except Exception as e:
        logging.error(f"Error deleting announcement: {e}")
        raise HTTPException(status_code=500, detail=f"Error deleting announcement: {str(e)}")

if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)