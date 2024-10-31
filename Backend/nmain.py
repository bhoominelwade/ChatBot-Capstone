# Required libraries for handling PDFs, images, and OCR
import fitz  # PyMuPDF for working with PDF files
import pytesseract  # OCR tool for extracting text from images
from PIL import Image  # Image processing library
import io  # For handling byte streams
import os  # For operating system dependent functionality

# Libraries for data manipulation and file handling
import pandas as pd  # Data analysis and manipulation library
from docx import Document as WordDocument  # For handling Word documents
from openpyxl import load_workbook  # For reading and writing Excel files
from datetime import datetime, timedelta
import mimetypes
from typing import Optional
import uuid
from firebase_admin import firestore

# LangChain specific libraries for embedding, text splitting, and vector storage
from langchain_community.embeddings import OllamaEmbeddings  # For embeddings
from langchain.text_splitter import RecursiveCharacterTextSplitter  # For splitting text
from langchain_community.vectorstores import Chroma  # For vector storage
from langchain.chains import ConversationalRetrievalChain  # For conversational chains
from langchain_community.chat_message_histories import ChatMessageHistory  # For message history management
from langchain.memory import ConversationBufferMemory  # For conversation memory

# For working with Groq API in LangChain
from langchain_groq import ChatGroq  # For integrating with Groq services

# Environment variable management
from dotenv import load_dotenv  # For loading environment variables from .env file

# FastAPI for building the API
from fastapi import FastAPI, UploadFile, File, HTTPException, Form  # For creating the FastAPI app and handling file uploads
from fastapi.middleware.cors import CORSMiddleware  # type: ignore # For handling CORS

# Uvicorn for serving the FastAPI app
import uvicorn  # ASGI server for running FastAPI applications

# Pydantic for data validation and settings management
from pydantic import BaseModel  # For defining data models

# Logging for application logging
import logging  # For logging messages and errors

# Firebase imports
import firebase_admin
from firebase_admin import credentials, storage, firestore
import datetime
from firebase_admin import storage

from tenacity import retry, stop_after_attempt, wait_exponential
from fuzzywuzzy import process

# Initialize Firebase
cred = credentials.Certificate(r"C:\Users\Ammar Abdulhussain\Desktop\phullstack\Backend\capstone-4eff9-firebase-adminsdk-a48fw-6daf6a9552.json")
firebase_admin.initialize_app(cred, {
    'storageBucket': 'capstone-4eff9.appspot.com'
})

# Get a reference to the storage service and Firestore
bucket = storage.bucket()
db = firestore.client()

# Loading environment variables from .env file
load_dotenv()

app = FastAPI()

logging.basicConfig(level=logging.INFO)

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Initialize conversation chain with GROQ language model
groq_api_key = os.environ['GROQ_API_KEY']

llm_groq = ChatGroq(
    groq_api_key=groq_api_key,
    model_name="llama-3.1-70b-versatile",
    temperature=0.2
)

# Global variable to store the conversation chain
global_chain = None

def extract_text_from_image(image_bytes):
    """
    Extract text from an image using OCR.
    """
    image = Image.open(io.BytesIO(image_bytes))
    ocr_text = pytesseract.image_to_string(image)
    return ocr_text

def normalize_role(role):
    """Normalize role names to lowercase singular form."""
    role = role.lower()
    if role.endswith('s'):
        role = role[:-1]  # Remove trailing 's'
    return role

def upload_to_firebase(file_content, filename, role):
    """
    Upload a file to Firebase Storage with metadata and correct MIME type.
    """
    blob = bucket.blob(filename)
    
    # Detect MIME type
    content_type, _ = mimetypes.guess_type(filename)
    if content_type is None:
        content_type = 'application/octet-stream'  # Default to binary if type can't be guessed
    
    blob.upload_from_string(file_content, content_type=content_type)
    
    # Normalize and set custom metadata
    normalized_role = normalize_role(role)
    metadata = {'role': normalized_role}
    blob.metadata = metadata
    blob.patch()

    # Verify that metadata was set correctly
    updated_blob = bucket.get_blob(filename)
    if updated_blob.metadata != metadata:
        logging.error(f"Failed to set metadata for {filename}. Expected {metadata}, got {updated_blob.metadata}")
    else:
        logging.info(f"File uploaded to Firebase with metadata: {filename}, role: {normalized_role}, content_type: {content_type}")

    return blob.public_url

def convert_timestamp(timestamp):
    if isinstance(timestamp, datetime.datetime):
        ist_time = timestamp + datetime.timedelta(hours=5, minutes=30)
        return ist_time.strftime('%Y-%m-%d %H:%M:%S')
    return None

def process_files(files, role):
    texts = []
    metadatas = []

    for file in files:
        file_content = file.file.read()
        file_ext = os.path.splitext(file.filename)[1].lower()
        logging.info(f"Processing file: {file.filename} with extension: {file_ext}")

        try:
            # Upload file to Firebase Storage with role metadata
            file_url = upload_to_firebase(file_content, file.filename, role)
            logging.info(f"File uploaded to Firebase: {file_url}")

            if file_ext in ['.pdf']:
                # Process PDF files
                pdf = fitz.open(stream=file_content, filetype="pdf")
                for page_num in range(len(pdf)):
                    page = pdf[page_num]
                    text = page.get_text()
                    if text:
                        texts.append(text)

                    images = page.get_images(full=True)
                    for img_index, img in enumerate(images, start=1):
                        xref = img[0]
                        base_image = pdf.extract_image(xref)
                        image_bytes = base_image["image"]

                        # OCR extraction
                        ocr_text = extract_text_from_image(image_bytes)
                        if ocr_text.strip():
                            texts.append(ocr_text)

            elif file_ext in ['.csv']:
                # Process CSV files
                df = pd.read_csv(io.StringIO(file_content.decode()))
                texts.append(df.to_string())

            elif file_ext in ['.txt']:
                # Process TXT files
                texts.append(file_content.decode())

            elif file_ext in ['.docx']:
                # Process Word files
                doc = WordDocument(io.BytesIO(file_content))
                text = "\n".join([para.text for para in doc.paragraphs])
                texts.append(text)

            elif file_ext in ['.xlsx']:
                # Process Excel files
                wb = load_workbook(io.BytesIO(file_content), data_only=True)
                text = ""
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for row in ws.iter_rows(values_only=True):
                        text += " ".join([str(cell) for cell in row]) + "\n"
                texts.append(text)

            elif file_ext in ['.jpeg', '.jpg', '.png']:
                # Process image files (JPEG, JPG, PNG)
                ocr_text = extract_text_from_image(file_content)
                if ocr_text.strip():
                    texts.append(ocr_text)

            else:
                logging.warning(f"Unsupported file type: {file_ext}")

        except Exception as e:
            logging.error(f"Error processing file {file.filename}: {e}")
            raise

    # Aggregate and split texts into chunks
    aggregated_text = " ".join(texts)
    text_splitter = RecursiveCharacterTextSplitter(chunk_size=1200, chunk_overlap=50)
    file_texts = text_splitter.split_text(aggregated_text)

    # Populate metadata for each chunk
    if file_texts:
        metadatas = [{"source": f"{i}-{file.filename}", "role": role} for i in range(len(file_texts))]

    if len(file_texts) == 0:
        raise ValueError("No texts extracted from the provided files.")

    # Create a Chroma vector store
    embeddings = OllamaEmbeddings(model="nomic-embed-text")
    docsearch = Chroma.from_texts(file_texts, embeddings, metadatas=metadatas)

    # Initialize message history for conversation
    message_history = ChatMessageHistory()

    # Memory for conversational context
    memory = ConversationBufferMemory(
        memory_key="chat_history",
        output_key="answer",
        chat_memory=message_history,
        return_messages=True,
    )

    # Create a chain that uses the Chroma vector store
    chain = ConversationalRetrievalChain.from_llm(
        llm=ChatGroq(),
        chain_type="stuff",
        retriever=docsearch.as_retriever(),
        memory=memory,
        return_source_documents=True,
    )

    return chain

def can_access_document(user_role, document_role):
    """
    Check if a user with given role can access a document with specified role.
    Implements role hierarchy: HOD/Dean > Teacher > Student
    """
    # Normalize roles to lowercase
    user_role = user_role.lower().replace('/', '').replace(' ', '')
    document_role = document_role.lower().replace('/', '').replace(' ', '')
    
    # Define role hierarchy
    role_hierarchy = {
        'hod_dean': ['hod_dean', 'teacher', 'student'],
        'teacher': ['teacher', 'student'],
        'student': ['student']
    }
    
    # Check if user's role exists in hierarchy
    if user_role not in role_hierarchy:
        logging.error(f"Invalid user role: {user_role}")
        return False
    
    # Check if document role is accessible to user
    return document_role in role_hierarchy[user_role]

async def process_timetable(file_content):
    """Process timetable file and extract schedule information"""
    try:
        wb = load_workbook(io.BytesIO(file_content))
        ws = wb.active
        schedule = {}
        
        # Skip header row
        for row in list(ws.rows)[1:]:
            time_slot = row[0].value
            if time_slot:  # Skip empty rows
                for col, day in enumerate(['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday'], 1):
                    if day not in schedule:
                        schedule[day] = {}
                    activity = row[col].value
                    if activity:  # Only add non-empty cells
                        schedule[day][time_slot] = str(activity).strip()
        
        return schedule
    except Exception as e:
        logging.error(f"Error processing timetable: {e}")
        raise

async def check_availability(teacher_id: str, date: str, time: str) -> tuple[bool, str]:
    """Check if a teacher is available at the specified date and time"""
    try:
        # Validate and format time
        try:
            # Handle various time formats
            if ':' not in time:
                if len(time) == 4:  # Format like "0900"
                    time = f"{time[:2]}:{time[2:]}"
                else:
                    return False, "Invalid time format. Please use HH:MM"
            
            hour = int(time.split(':')[0])
            if hour < 9 or hour >= 17:
                return False, "Invalid time. Available hours are between 09:00-17:00"
            
            # Format time slot
            time_slot = f"{time.zfill(5)}-{str(hour+1).zfill(2)}:00"
            
        except (ValueError, IndexError):
            return False, "Invalid time format. Please use HH:MM"

        # Validate date format and check weekend
        try:
            requested_date = datetime.datetime.strptime(date, '%Y-%m-%d')
            day_of_week = requested_date.strftime('%A')
            if day_of_week in ['Saturday', 'Sunday']:
                return False, "Meetings can only be scheduled on weekdays"
        except ValueError:
            return False, "Invalid date format. Please use YYYY-MM-DD"

        # Check lunch break
        if hour == 13:
            return False, "This is lunch break time (13:00-14:00)"

        # Get timetable from Firestore
        timetable_ref = db.collection('timetables').document(teacher_id)
        timetable_doc = timetable_ref.get()

        if timetable_doc.exists:
            timetable_data = timetable_doc.to_dict()
            if day_of_week in timetable_data:
                for slot, activity in timetable_data[day_of_week].items():
                    if slot == time_slot:
                        if activity == 'Office Hours':
                            return True, "Available during Office Hours"
                        if activity == 'Lunch Break':
                            return False, "This is lunch break time"
                        return False, f"Teacher has {activity} at this time"

        # Check existing meetings
        meetings_ref = db.collection('meetings')
        existing_meetings = meetings_ref.where('teacher_id', '==', teacher_id)\
                                     .where('date', '==', date)\
                                     .where('time', '==', time)\
                                     .get()
        
        if len(list(existing_meetings)) > 0:
            return False, "This time slot is already booked"

        return True, "Available"

    except Exception as e:
        logging.error(f"Error checking availability: {str(e)}")
        return False, f"Error checking availability: {str(e)}"

async def get_available_slots(teacher_id: str, date: str) -> list[str]:
    """Get available time slots for a teacher on a given date"""
    try:
        requested_date = datetime.datetime.strptime(date, '%Y-%m-%d')
        day_of_week = requested_date.strftime('%A')
        
        if requested_date.date() < datetime.datetime.now().date():
            return []
        
        if requested_date.weekday() >= 5:  # Weekend check
            return []

        # Define all possible slots excluding lunch hour
        all_slots = [f"{str(hour).zfill(2)}:00" for hour in range(9, 17) if hour != 13]
        available_slots = []
        
        # Get teacher's timetable
        timetable_ref = db.collection('timetables').document(teacher_id)
        timetable_doc = timetable_ref.get()
        
        if timetable_doc.exists:
            timetable_data = timetable_doc.to_dict()
            
            for time in all_slots:
                hour = int(time.split(':')[0])
                time_slot = f"{time}-{str(hour+1).zfill(2)}:00"
                
                # Check if slot is free in timetable
                if day_of_week in timetable_data:
                    activity = timetable_data[day_of_week].get(time_slot)
                    if not activity or activity == 'Office Hours':
                        # Check existing meetings
                        meetings_ref = db.collection('meetings')
                        existing_meetings = meetings_ref.where('teacher_id', '==', teacher_id)\
                                                     .where('date', '==', date)\
                                                     .where('time', '==', time)\
                                                     .get()
                        
                        if len(list(existing_meetings)) == 0:
                            available_slots.append(time)

        return available_slots
    except Exception as e:
        logging.error(f"Error getting available slots: {str(e)}")
        return []

@app.post("/upload/")
async def upload_file(files: list[UploadFile] = File(...), role: str = Form(...)):
    global global_chain
    try:
        normalized_role = normalize_role(role)
        logging.info(f"Received {len(files)} files for processing with normalized role: {normalized_role}")
        global_chain = process_files(files, normalized_role)
        logging.info("Files processed successfully.")
        return {"message": f"Successfully processed {len(files)} files for role: {normalized_role}"}
    except Exception as e:
        logging.error(f"Error processing files: {e}")
        raise HTTPException(status_code=500, detail=str(e))

class ChatRequest(BaseModel):
    message: str

@app.post("/retrieve-document/")
async def retrieve_document(request: dict):
    file_name = request.get("fileName")
    user_role = normalize_role(request.get("userRole", ""))

    logging.info(f"Document retrieval request - File: {file_name}, User Role: {user_role}")

    if not file_name or not user_role:
        logging.error("Missing fileName or userRole in request")
        raise HTTPException(status_code=400, detail="Missing fileName or userRole")

    try:
        # Get list of all blobs in the bucket
        blobs = list(bucket.list_blobs())
        blob_names = [blob.name for blob in blobs]

        # Find the best matching file
        best_match, score = process.extractOne(file_name, blob_names)
        if score < 80:
            logging.error(f"No close match found for: {file_name}")
            raise HTTPException(status_code=404, detail=f"Document not found: {file_name}")

        blob = bucket.blob(best_match)
        
        # Get document metadata
        blob.reload()  # Ensure we have the latest metadata
        document_role = blob.metadata.get('role', 'student') if blob.metadata else 'student'
        
        logging.info(f"Document metadata - File Role: {document_role}")
        
        # Check access permission
        if not can_access_document(user_role, document_role):
            logging.warning(f"Access denied - User Role: {user_role}, Document Role: {document_role}")
            return {
                "canAccess": False,
                "message": "You don't have permission to access this document."
            }

        # Generate signed URL for authorized access
        signed_url = blob.generate_signed_url(
            version="v4",
            expiration=datetime.timedelta(minutes=15),
            method="GET"
        )
        
        logging.info(f"Access granted - Generated signed URL for {best_match}")
        return {
            "canAccess": True,
            "downloadUrl": signed_url,
            "fileName": best_match
        }

    except Exception as e:
        logging.exception(f"Error in retrieve_document: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")
    
@app.post("/chat/")
async def chat(request: ChatRequest):
    """
    Simplified chat endpoint that uses a single global chain for all users
    """
    global global_chain
    
    if not global_chain:
        raise HTTPException(
            status_code=400, 
            detail="No documents have been processed for chat. Please ensure documents are uploaded first."
        )

    try:
        prompt = (
            "You are a friendly and professional university chatbot. "
            "Your tone should be polite and engaging, suitable for a teenage audience. "
            "You should provide helpful responses related to the university documents and general inquiries. "
            "Feel free to offer casual conversation while staying within a professional and respectful tone. "
            "Answer to the question in a clear and concise manner. "
            "Also respond to general greetings and notes in a friendly, conversational tone. "
            "Respond to thanking and similar statements with welcoming notes. "
        )

        context = f"{prompt}\nUser: {request.message}\nBot:"
        
        logging.info(f"Processing chat request with message: {request.message[:100]}...")
        
        res = global_chain.invoke({"question": context})
        answer = res["answer"]
        source_documents = res["source_documents"]

        if source_documents:
            source_names = [f"source_{i}" for i in range(len(source_documents))]
            answer += f"\nSources: {', '.join(source_names)}"

        logging.info("Chat request processed successfully")
        return {"response": answer}

    except Exception as e:
        error_msg = f"Error processing chat request: {str(e)}"
        logging.error(error_msg)
        raise HTTPException(status_code=500, detail=error_msg)
    
@app.post("/announcement/")
async def create_announcement(announcement: dict):
    try:
        announcement_ref = db.collection('announcements').document()
        
        announcement_data = {
            'title': announcement.get('title'),
            'text': announcement.get('text'),
            'role': normalize_role(announcement.get('role')),
            'isImportant': announcement.get('isImportant', False),
            'timestamp': firestore.SERVER_TIMESTAMP,
            'formattedText': True
        }
        
        announcement_ref.set(announcement_data)
        
        return {
            "status": "success",
            "message": "Announcement created successfully",
            "id": announcement_ref.id
        }
    except Exception as e:
        logging.error(f"Error creating announcement: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    

@app.get("/announcements/{role}")
async def get_announcements(role: str, limit: int = 5):
    try:
        # Normalize role
        role = normalize_role(role)
        
        # Get Firestore database reference
        announcements_ref = db.collection('announcements')\
            .order_by('timestamp', direction=firestore.Query.DESCENDING)
        
        docs = announcements_ref.get()
        
        announcements = []
        for doc in docs:
            data = doc.to_dict()
            data['id'] = doc.id
            
            if data.get('timestamp'):
                utc_time = data['timestamp']
                if isinstance(utc_time, datetime.datetime):
                    ist_time = utc_time + datetime.timedelta(hours=5, minutes=30)
                    data['timestamp'] = ist_time.strftime('%d/%m/%Y %H:%M')
            
            doc_role = normalize_role(data.get('role', ''))
            should_include = False
            
            if role == 'hod_dean':
                should_include = True
            elif role == 'teacher' and doc_role in ['teacher', 'student']:
                should_include = True
            elif role == 'student' and doc_role == 'student':
                should_include = True
            
            if should_include:
                announcements.append(data)
        
        return {"announcements": announcements[:limit]}
            
    except Exception as e:
        logging.error(f"Error fetching announcements: {e}")
        return {"announcements": []}


@app.delete("/announcement/{announcement_id}")
async def delete_announcement(announcement_id: str):
    """Delete an announcement by ID."""
    try:
        # Get Firestore database reference
        db = firestore.client()
        
        # Delete the announcement document
        db.collection('announcements').document(announcement_id).delete()
        
        return {"message": "Announcement deleted successfully", "status": "success"}
    except Exception as e:
        logging.error(f"Error deleting announcement: {e}")
        raise HTTPException(status_code=500, detail=f"Error deleting announcement: {str(e)}")

@app.post("/upload-timetable/")
async def upload_timetable(file: UploadFile = File(...), teacherId: str = Form(...)):
    """Upload a teacher's timetable"""
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Only Excel files are allowed")

    try:
        file_content = await file.read()
        
        # Process timetable to validate format
        schedule = await process_timetable(file_content)
        logging.info(f"Processed schedule: {schedule}")  # Add logging
        
        # Upload to Firebase Storage with consistent naming
        storage_filename = f"timetables/{teacherId}_timetable.xlsx"
        blob = bucket.blob(storage_filename)
        blob.upload_from_string(
            file_content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Store processed schedule in Firestore with consistent ID
        timetable_ref = db.collection('timetables').document(teacherId)
        timetable_ref.set(schedule)
        
        logging.info(f"Uploaded timetable for teacher {teacherId}")
        logging.info(f"Storage path: {storage_filename}")
        logging.info(f"Firestore collection: timetables/{teacherId}")

        return {
            "message": "Timetable uploaded successfully",
            "storage_path": storage_filename,
            "schedule": schedule  # Return the processed schedule for verification
        }
    except Exception as e:
        logging.error(f"Error uploading timetable: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/book-meeting/")
async def create_meeting(
    teacher_name: str = Form(...),
    student_id: str = Form(...),
    date: str = Form(...),
    time: str = Form(...)
):
    """Create a new meeting with proper validation"""
    try:
        # First find teacher by name
        users_ref = db.collection('Users')
        teacher_query = users_ref.where('name', '==', teacher_name).where('role', 'in', ['teacher', 'hod_dean']).limit(1).get()
        
        if not teacher_query:
            return {
                "success": False,
                "message": f"Teacher '{teacher_name}' not found",
                "available_slots": []
            }

        teacher_id = teacher_query[0].id

        # Validate date and time format
        try:
            requested_date = datetime.datetime.strptime(date, '%Y-%m-%d')
            # Check if date is in the past
            if requested_date.date() < datetime.datetime.now().date():
                return {
                    "success": False,
                    "message": "Cannot book meetings for past dates",
                    "available_slots": []
                }
            
            # Check if it's a weekend
            if requested_date.weekday() >= 5:  # 5 is Saturday, 6 is Sunday
                return {
                    "success": False,
                    "message": "Cannot book meetings on weekends",
                    "available_slots": []
                }

            # Format time and validate
            if ':' not in time:
                if len(time) == 4:  # Format like "0900"
                    time = f"{time[:2]}:{time[2:]}"
                else:
                    return {
                        "success": False,
                        "message": "Invalid time format. Please use HH:MM",
                        "available_slots": []
                    }
            
            hour = int(time.split(':')[0])
            if hour < 9 or hour >= 17:
                return {
                    "success": False,
                    "message": "Invalid time. Available hours are between 09:00-17:00",
                    "available_slots": []
                }

            # Check lunch break
            if hour == 13:
                return {
                    "success": False,
                    "message": "Cannot book during lunch break (13:00-14:00)",
                    "available_slots": []
                }

        except ValueError:
            return {
                "success": False,
                "message": "Invalid date or time format",
                "available_slots": []
            }

        # Check timetable conflicts
        day_of_week = requested_date.strftime('%A')
        time_slot = f"{time}-{str(hour+1).zfill(2)}:00"

        # Get teacher's timetable
        timetable_ref = db.collection('timetables').document(teacher_id)
        timetable_doc = timetable_ref.get()

        if timetable_doc.exists:
            timetable_data = timetable_doc.to_dict()
            if day_of_week in timetable_data:
                activity = timetable_data[day_of_week].get(time_slot)
                if activity and activity != 'Office Hours':
                    return {
                        "success": False,
                        "message": f"Teacher has {activity} at this time",
                        "available_slots": await get_available_slots(teacher_id, date)
                    }

        # Check existing meetings
        meetings_ref = db.collection('meetings')
        existing_meetings = meetings_ref.where('teacher_id', '==', teacher_id)\
                                     .where('date', '==', date)\
                                     .where('time', '==', time)\
                                     .get()
        
        if len(list(existing_meetings)) > 0:
            return {
                "success": False,
                "message": "This time slot is already booked",
                "available_slots": await get_available_slots(teacher_id, date)
            }

        # If all validations pass, create the meeting
        meeting_ref = db.collection('meetings').document()
        meeting_data = {
            'teacher_id': teacher_id,
            'teacher_name': teacher_name,
            'student_id': student_id,
            'date': date,
            'time': time,
            'status': 'scheduled',
            'created_at': firestore.SERVER_TIMESTAMP
        }
        
        meeting_ref.set(meeting_data)
        
        return {
            "success": True,
            "message": "Meeting scheduled successfully",
            "meeting_id": meeting_ref.id
        }

    except Exception as e:
        logging.error(f"Error creating meeting: {str(e)}")
        return {
            "success": False,
            "message": f"Failed to schedule meeting: {str(e)}",
            "available_slots": []
        }

@app.get("/available-slots/{teacher_id}/{date}")
async def get_teacher_slots(teacher_id: str, date: str):
    """Get available meeting slots for a teacher on a specific date"""
    try:
        slots = await get_available_slots(teacher_id, date)
        return {"available_slots": slots}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/teachers/")
async def get_teachers():
    """Get list of teachers and HOD/Deans"""
    try:
        # Get Firestore database reference
        users_ref = db.collection('Users')
        
        # Create query for teachers and HOD/Deans
        query = users_ref.where('role', 'in', ['teacher', 'hod_dean'])
        teachers_snapshot = query.get()
        
        # Process results
        teacher_list = []
        for doc in teachers_snapshot:
            data = doc.to_dict()
            # Add required fields
            teacher_data = {
                'id': doc.id,
                'name': data.get('name', 'Unknown'),
                'role': data.get('role', 'teacher'),
                'email': data.get('email', ''),
                # Add any other fields you need
            }
            teacher_list.append(teacher_data)
            
        logging.info(f"Successfully fetched {len(teacher_list)} teachers")
        return {"teachers": teacher_list}
        
    except Exception as e:
        logging.error(f"Error fetching teachers: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to fetch teachers: {str(e)}"
        )

@app.get("/timetable/{teacher_id}")
async def get_timetable(teacher_id: str):
    """Get a teacher's timetable"""
    try:
        # Get timetable from Firestore
        timetable_ref = db.collection('timetables').document(teacher_id)
        timetable_doc = timetable_ref.get()
        
        if not timetable_doc.exists:
            return {
                "success": False,
                "message": "No timetable found for this teacher"
            }
            
        return {
            "success": True,
            "timetable": timetable_doc.to_dict()
        }
    except Exception as e:
        logging.error(f"Error fetching timetable: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/meeting/{meeting_id}")
async def delete_meeting(meeting_id: str):
    """Delete a meeting"""
    try:
        meeting_ref = db.collection('meetings').document(meeting_id)
        meeting_ref.delete()
        return {"message": "Meeting deleted successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/meetings/{user_id}")
async def get_meetings(user_id: str, user_role: str = None):
    """
    Get meetings for a user (either teacher or student)
    
    Args:
        user_id: The ID of the user (teacher or student)
        user_role: Optional role parameter to optimize query
    """
    try:
        meetings_ref = db.collection('meetings')
        meetings = []

        # If role is specified, optimize the query
        if user_role:
            if user_role.lower() in ['teacher', 'hod_dean']:
                meetings_query = meetings_ref.where('teacher_id', '==', user_id)
            else:
                meetings_query = meetings_ref.where('student_id', '==', user_id)
        else:
            # If no role specified, check both teacher and student meetings
            teacher_meetings = meetings_ref.where('teacher_id', '==', user_id).get()
            student_meetings = meetings_ref.where('student_id', '==', user_id).get()
            
            # Combine both queries
            meetings_query = list(teacher_meetings) + list(student_meetings)

        # If we have a query (role was specified), execute it
        if isinstance(meetings_query, firestore.Query):
            meetings_query = meetings_query.get()

        # Process meetings
        seen_ids = set()  # To prevent duplicates
        for meeting in meetings_query:
            meeting_id = meeting.id
            if meeting_id not in seen_ids:
                seen_ids.add(meeting_id)
                data = meeting.to_dict()
                data['id'] = meeting_id
                
                # Convert timestamps to strings
                if isinstance(data.get('created_at'), datetime.datetime):
                    ist_time = data['created_at'] + datetime.timedelta(hours=5, minutes=30)
                    data['created_at'] = ist_time.strftime('%Y-%m-%d %H:%M:%S')
                
                # Ensure date is in correct format
                if isinstance(data.get('date'), datetime.datetime):
                    data['date'] = data['date'].strftime('%Y-%m-%d')
                
                meetings.append(data)

        # Sort meetings by date and time
        meetings.sort(key=lambda x: (x['date'], x['time']))
        
        logging.info(f"Retrieved {len(meetings)} meetings for user {user_id}")
        return {
            "success": True,
            "meetings": meetings,
            "count": len(meetings)
        }
    except Exception as e:
        error_msg = f"Error fetching meetings for user {user_id}: {str(e)}"
        logging.error(error_msg)
        return {
            "success": False,
            "message": error_msg,
            "meetings": [],
            "count": 0
        }

        
@app.get("/recent-files/{role}")
async def get_recent_files(role: str):
    """Get 8 most recent files for a role."""
    try:
        # Get all files from storage bucket
        blobs = list(bucket.list_blobs())
        
        # Filter and sort files
        role_files = []
        for blob in blobs:
            if blob.metadata and blob.metadata.get('role') == normalize_role(role):
                file_info = {
                    'name': blob.name,
                    'size': blob.size,
                    'uploaded': blob.time_created.strftime('%Y-%m-%d %H:%M:%S'),
                    'role': normalize_role(role),
                    'type': blob.content_type,
                }
                role_files.append(file_info)
        
        # Sort by upload time (newest first) and limit to 8
        role_files.sort(key=lambda x: x['uploaded'], reverse=True)
        recent_files = role_files[:8]
        
        return {"files": recent_files}
    except Exception as e:
        logging.error(f"Error fetching recent files: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
    
@app.delete("/delete-file/{file_name:path}")
async def delete_file(file_name: str):
    """
    Delete a file from Firebase Storage
    
    Args:
        file_name: The name/path of the file to delete
    """
    try:
        logging.info(f"Attempting to delete file: {file_name}")
        
        # Get the blob reference
        blob = bucket.blob(file_name)
        
        # Check if blob exists
        if not blob.exists():
            logging.error(f"File not found: {file_name}")
            raise HTTPException(
                status_code=404,
                detail=f"File not found: {file_name}"
            )
            
        # Delete the blob
        blob.delete()
        logging.info(f"Successfully deleted file: {file_name}")
        
        # Also clear the file from the chat system if it was used
        global global_chain
        if global_chain:
            # Reset the global chain to force reprocessing of remaining files
            global_chain = None
            
        return {
            "success": True,
            "message": f"Successfully deleted file: {file_name}"
        }
        
    except Exception as e:
        error_msg = f"Error deleting file {file_name}: {str(e)}"
        logging.error(error_msg)
        raise HTTPException(
            status_code=500,
            detail=error_msg
        )

if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)